#!/usr/bin/env python3
"""
Build data.json for the Micro Demand per-doctor dashboard from the source tabs of
the "Central team - Micro Demand processing dashboard" workbook.

Usage:
    python3 build_data.py [path-to-xlsx] [output-json]
Defaults: ./source/Micro_Demand.xlsx -> ./data.json

Source tabs (all daily grain, city / locality(clinic) / doctor):
  RD-Rev        dt,city,locality,provider,channel,users,payments
  RD-Utliliz    dt,city,locality,doctor, + SC/FU slot & time columns
  RD-Avail      dt,city,locality,doctor, roster/shrinkage/net mins, weekend-weekday
  RD - Bookings and done   appt-level counts (offline/online/repeat booked/done/noshow/reschedule + task funnel)
  RD - D2P      appt-level completed/purchased/prescribe values (online/offline/repeat)
  RD - Pat level B2D        patient-level (for unique Booking(P)/Done(P) counts)

Everything is aggregated to (day, city, clinic, doctor); the HTML sums over the
selected City/Clinic/Doctor filter and the month/week/day date buckets, and
computes ratios from the summed components — reproducing the sheet's SUMIFS exactly
and generalising to any filter level (Excel <>criteria = match-all).
"""
import sys, os, json
from collections import defaultdict
from datetime import date
import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
XLSX = sys.argv[1] if len(sys.argv) > 1 else os.path.join(HERE, 'source', 'Micro_Demand.xlsx')
OUT = sys.argv[2] if len(sys.argv) > 2 else os.path.join(HERE, 'data.json')

wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)

# ---- shared dimension registries (display name + casefold match, Excel-style) ----
cities, clinics, doctors, channels, patients, diags = {}, {}, {}, {}, {}, {}
def reg(d, name):
    key = (name or '').strip()
    if key not in d:
        d[key] = len(d)
    return d[key]

def dord(v):
    return v.toordinal() if hasattr(v, 'toordinal') else None

minday = [None]
def dayoff(v):
    o = dord(v)
    if o is None:
        return None
    if minday[0] is None or o < minday[0]:
        minday[0] = o
    return o

def num(v):
    return v if isinstance(v, (int, float)) else 0

# We stream twice conceptually: collect raw with absolute ordinals, then rebase.
rev = defaultdict(float)              # (ord,ci,cl,doc,ch) -> pay
util = defaultdict(lambda: [0.0]*10)  # (ord,ci,cl,doc) -> [E,F,H,J,L,N,O,P,R,T]
avail = {}                            # (ord,ci,cl,doc) -> [roster,shrink,netSC,netRpt, wknd(0/1)]
book = defaultdict(lambda: [0.0]*36)  # (ord,ci,cl,doc) -> cols E..AN (idx4..39)
d2p = defaultdict(lambda: [0.0]*41)   # (ord,ci,cl,doc,diag) -> cols F..AT (idx5..45)
pat = []                              # [ord,ci,cl,doc,patI, typeFlag(0=SC,1=FU), offlineFlag, completed]

# ---- RD-Rev ----
sh = wb['RD-Rev']
for r in sh.iter_rows(min_row=2, values_only=True):
    if not r or r[0] is None: continue
    o = dayoff(r[0])
    k = (o, reg(cities, r[1]), reg(clinics, r[2]), reg(doctors, r[3]), reg(channels, r[4]))
    rev[k] += num(r[6])

# ---- RD-Utliliz ----  need E,F,H,J,L,N,O,P,R,T -> src idx 4,5,7,9,11,13,14,15,17,19
UCOLS = [4,5,7,9,11,13,14,15,17,19]
sh = wb['RD-Utliliz']
for r in sh.iter_rows(min_row=2, values_only=True):
    if not r or r[0] is None: continue
    o = dayoff(r[0])
    k = (o, reg(cities, r[1]), reg(clinics, r[2]), reg(doctors, r[3]))
    a = util[k]
    for i, ci in enumerate(UCOLS):
        a[i] += num(r[ci]) if ci < len(r) else 0

# ---- RD-Avail ----  E=roster(4) F=shrink(5) G=netSC(6) J=netRpt(9) M=weekend/weekday(12)
sh = wb['RD-Avail']
for r in sh.iter_rows(min_row=2, values_only=True):
    if not r or r[0] is None: continue
    o = dayoff(r[0])
    k = (o, reg(cities, r[1]), reg(clinics, r[2]), reg(doctors, r[3]))
    wl = str(r[12]).strip().lower() if len(r) > 12 and r[12] is not None else ''
    wknd = 1 if wl.startswith('weekend') else 0
    cur = avail.get(k)
    add = [num(r[4]), num(r[5]), num(r[6]), num(r[9]), wknd]
    if cur is None:
        avail[k] = add
    else:
        for i in range(4): cur[i] += add[i]
        cur[4] = wknd

# ---- RD - Bookings and done ----  cols E..AN = idx 4..39
sh = wb['RD - Bookings and done']
for r in sh.iter_rows(min_row=2, values_only=True):
    if not r or r[0] is None: continue
    o = dayoff(r[0])
    k = (o, reg(cities, r[1]), reg(clinics, r[2]), reg(doctors, r[3]))
    a = book[k]
    for i in range(36):
        ci = 4 + i
        if ci < len(r): a[i] += num(r[ci])

# ---- RD - D2P ----  cols F..AT = idx 5..45 (diagnosis col E idx4 ignored: <> wildcard)
sh = wb['RD - D2P']
for r in sh.iter_rows(min_row=2, values_only=True):
    if not r or r[0] is None: continue
    o = dayoff(r[0])
    dg = reg(diags, r[4]) if len(r) > 4 else reg(diags, '')
    k = (o, reg(cities, r[1]), reg(clinics, r[2]), reg(doctors, r[3]), dg)
    a = d2p[k]
    for i in range(41):
        ci = 5 + i
        if ci < len(r): a[i] += num(r[ci])

# ---- RD - Pat level B2D ----  A=dt,B=patient,D=doctor,E=type,F=city,G=clinic,H=status,J=offline_flag
sh = wb['RD - Pat level B2D']
for r in sh.iter_rows(min_row=2, values_only=True):
    if not r or r[0] is None: continue
    typ = str(r[4]).strip() if len(r) > 4 and r[4] is not None else ''
    tl = typ.lower()
    if tl == 'screening call': tf = 0
    elif tl == 'follow up': tf = 1
    else: continue  # only SC + FU feed the tabs
    o = dayoff(r[0])
    ci = reg(cities, r[5]); cl = reg(clinics, r[6]); doc = reg(doctors, r[3])
    pt = reg(patients, r[1])
    off = 1 if (len(r) > 9 and r[9] in (1, 1.0, '1', True)) else 0
    comp = 1 if (len(r) > 7 and str(r[7]).strip().upper() == 'COMPLETED') else 0
    pat.append((o, ci, cl, doc, pt, tf, off, comp))

wb.close()

base = minday[0]
maxoff = 0
def rb(o):
    global maxoff
    v = o - base
    if v > maxoff: maxoff = v
    return v

data = {
    'cities': list(cities.keys()),
    'clinics': list(clinics.keys()),
    'doctors': list(doctors.keys()),
    'channels': list(channels.keys()),
    'diagnoses': list(diags.keys()),
    'rev':   [[rb(k[0]), k[1], k[2], k[3], k[4], round(v, 2)] for k, v in rev.items()],
    'util':  [[rb(k[0]), k[1], k[2], k[3]] + [round(x, 2) for x in v] for k, v in util.items()],
    'avail': [[rb(k[0]), k[1], k[2], k[3]] + [round(v[0], 2), round(v[1], 2), round(v[2], 2), round(v[3], 2), v[4]] for k, v in avail.items()],
    'book':  [[rb(k[0]), k[1], k[2], k[3]] + [round(x, 2) for x in v] for k, v in book.items()],
    'd2p':   [[rb(k[0]), k[1], k[2], k[3], k[4]] + [round(x, 2) for x in v] for k, v in d2p.items()],
    'pat':   [[rb(p[0]), p[1], p[2], p[3], p[4], p[5], p[6], p[7]] for p in pat],
}
bd = date.fromordinal(base)
md = date.fromordinal(base + maxoff)
data['base_date'] = [bd.year, bd.month, bd.day]
data['data_max'] = [md.year, md.month, md.day]
data['span_days'] = maxoff + 1

with open(OUT, 'w') as f:
    json.dump(data, f, separators=(',', ':'))
print('%s  %.2fMB' % (OUT, os.path.getsize(OUT)/1e6))
print('cities=%d clinics=%d doctors=%d channels=%s' % (len(cities), len(clinics), len(doctors), list(channels.keys())))
print('rev=%d util=%d avail=%d book=%d d2p=%d pat=%d  base=%s max=%s' %
      (len(rev), len(util), len(avail), len(book), len(d2p), len(pat), bd, md))
